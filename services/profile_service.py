"""
客户研判模块 — 业务服务层
===========================================
客户画像研判相关的所有业务逻辑。

核心职责:
  1. 客户资料上传    — 接收文本/文件，创建 customer_source 记录
  2. 客户来源查询    — 分页列表查询
  3. 研判结果查询    — 联合查询 customer_source + customer_profile
  4. AI 研判触发     — 根据画像规则，调用 AI 进行研判
  5. 画像规则管理    — CRUD（创建/列表/更新）

数据流（对应需求文档 4.1 客户研判模块）:
  客户信息进入 → customer_source（记录来源）→ AI 解析
              → 读取 profile_rule（画像规则）
              → 调用 Dify AI 进行研判
              → 写入 customer_profile（研判结果）

无物理外键策略:
  - operator_id → 校验 sys_user 存在
  - source_id   → 校验 customer_source 存在
  - evaluator_id → 校验 sys_user 存在

异步流程（对应 API 文档第 11 章 + 需求文档 CR-003）:
  研判是异步任务——上传后立即返回 source_id，
  AI 研判在后台执行，前端通过 GET /profile/{source_id} 轮询状态。

事务边界:
  - 上传/创建操作在请求事务中完成
  - AI 调用使用独立 Session（不在请求事务中调用外部 API）

参考文档:
  《教育服务系统_API接口设计规范文档_V1.2》
  - 第 9 章   客户研判模块接口
  - 第 11 章  异步任务接口规范
  - 第 14 章  应用层数据一致性保障
  《教育服务系统_需求规格说明书》
  - 第 4.1 节 客户研判模块
"""

import json
import logging
import os
from datetime import datetime
from typing import Any, Dict, List, Optional

from sqlalchemy.orm import Session

# --- 规则文件路径 ---
from config import PRODUCT_RULES_PATH, PRODUCT_CATALOG_PATH

# --- 公共基础 ---
from models.common import (
    ConflictError,
    NotFoundError,
    ReferenceNotFoundError,
    StateError,
    ValidationError,
    success_response,
)

# --- 数据模型 ---
from models.crm import CustomerProfile, CustomerSource, ProfileRule
from models.user import SysUser

# --- Schema ---
from schemas import PaginationParams
from schemas.crm import (
    CustomerProfileResponse,
    CustomerSourceResponse,
    ProfileDetailResponse,
    ProfileRuleCreate,
    ProfileRuleResponse,
    ProfileRuleUpdate,
)

# --- 数据库会话（异步任务使用独立 Session）---
from utils.database import SessionLocal

logger = logging.getLogger(__name__)


# ============================================================
# 一、客户资料上传
# ============================================================

def upload_customer_source(
    db: Session,
    source_type: str,
    operator_id: Optional[int] = None,
    content_text: Optional[str] = None,
    file_url: Optional[str] = None,
    file_name: Optional[str] = None,
) -> CustomerSource:
    """
    上传/录入客户资料。

    对应 API 文档第 9.2 节 POST /api/v1/profile/upload。

    业务规则:
      1. content_text 和 file_url 至少提供一个
      2. 如果传了 operator_id，校验操作人存在（无物理外键策略）
      3. 新建记录 parse_status = 'pending'

    参数:
        db:           数据库会话
        source_type:  来源类型（text/pdf_resume/excel/import/manual）
        operator_id:  操作人 ID（逻辑关联 sys_user）
        content_text: 文本内容
        file_url:     上传文件路径
        file_name:    原始文件名

    返回:
        新创建的 CustomerSource ORM 对象
    """
    # 1. 参数校验
    if not content_text and not file_url:
        raise ValidationError("content_text 和 file_url 至少提供一个")

    # 2. ⭐ 逻辑外键校验：操作人存在
    if operator_id is not None:
        operator = db.query(SysUser).filter_by(id=operator_id).first()
        if not operator:
            raise ReferenceNotFoundError("操作人", operator_id)

    # 3. 创建记录
    source = CustomerSource(
        source_type=source_type,
        raw_content=content_text,
        file_url=file_url,
        file_name=file_name,
        parse_status="pending",  # 新建记录等待解析
        operator_id=operator_id,
    )
    db.add(source)
    db.commit()
    db.refresh(source)

    return source


# ============================================================
# 二、客户来源查询
# ============================================================

def list_customer_sources(
    db: Session,
    pagination: PaginationParams,
    source_type: Optional[str] = None,
    parse_status: Optional[str] = None,
    operator_id: Optional[int] = None,
) -> Dict[str, Any]:
    """
    分页查询客户来源列表。

    支持筛选:
      - source_type:  按来源类型过滤
      - parse_status: 按解析状态过滤（pending/success/failed）
      - operator_id:  按操作人过滤

    返回:
        {items: [...], total: int, page: int, page_size: int}
    """
    query = db.query(CustomerSource)

    # 应用可选筛选条件
    if source_type:
        query = query.filter(CustomerSource.source_type == source_type)
    if parse_status:
        query = query.filter(CustomerSource.parse_status == parse_status)
    if operator_id is not None:
        query = query.filter(CustomerSource.operator_id == operator_id)

    total = query.count()
    sources = (
        query.order_by(CustomerSource.create_time.desc())
        .offset(pagination.skip)
        .limit(pagination.limit)
        .all()
    )

    items = [CustomerSourceResponse.model_validate(s).model_dump() for s in sources]
    return {
        "items": items,
        "total": total,
        "page": pagination.page,
        "page_size": pagination.page_size,
    }


# ============================================================
# 三、研判结果查询（联合查询）
# ============================================================


def get_profile_detail(db: Session, source_id: int) -> ProfileDetailResponse:
    """
    查询客户研判完整详情（来源 + 画像联合查询）。

    对应 API 文档第 9.4 节 GET /api/v1/profile/{source_id}。

    查询逻辑:
      1. 查 customer_source（来源记录）
      2. 查 customer_profile（通过 source_id 逻辑关联）
      3. 合并为 ProfileDetailResponse 返回

    如果研判尚未完成（customer_profile 不存在），画像字段全部为 null。
    前端根据 parse_status 判断是"等待中"还是"已完成"。

    参数:
        db:        数据库会话
        source_id: 来源记录 ID

    返回:
        ProfileDetailResponse（合并了 source + profile 的完整信息）
    """
    # 1. 查来源记录
    source = db.query(CustomerSource).filter_by(id=source_id).first()
    if not source:
        raise NotFoundError(f"客户来源记录不存在: id={source_id}")

    # 2. 查画像记录（通过 source_id 逻辑关联）
    profile = (
        db.query(CustomerProfile)
        .filter_by(source_id=source_id)
        .order_by(CustomerProfile.create_time.desc())  # 最新研判排在前面
        .first()
    )

    # 3. 合并返回
    return ProfileDetailResponse(
        # 来源信息
        source_id=source.id,
        source_type=source.source_type,
        file_url=source.file_url,
        file_name=source.file_name,
        parse_status=source.parse_status,
        parse_error=source.parse_error,
        # 画像信息（研判未完成时为 None）
        customer_name=profile.customer_name if profile else None,
        contact_info=profile.contact_info if profile else None,
        background_info=profile.background_info if profile else None,
        match_result=profile.match_result if profile else None,
        matched_product=profile.matched_product if profile else None,
        match_score=profile.match_score if profile else None,
        match_reason=profile.match_reason if profile else None,
        recommended_programs=profile.recommended_programs if profile else None,
        # 时间
        source_create_time=source.create_time,
        profile_update_time=profile.update_time if profile else None,
    )


# ============================================================
# 四、AI 研判触发（异步）
# ============================================================


def trigger_analysis(
    db: Session,
    source_id: int,
    evaluator_id: Optional[int] = None,
    rule_id: Optional[int] = None,
) -> Dict[str, Any]:
    """
    触发 AI 研判任务。

    对应 API 文档第 9.3 节 POST /api/v1/profile/{source_id}/analyze。

    流程:
      1. 校验来源记录存在且 parse_status 允许研判
      2. 记录 evaluator_id（可为系统自动）
      3. 提交后台任务（异步执行 AI 调用）
      4. 立即返回 source_id + 当前状态

    ⚠️ 异步执行约束（API 文档第 14.5 节）:
      - 请求事务中只做: 校验 + 更新状态为 pending + 提交后台任务
      - AI 调用在后台独立 Session 中执行
      - 不在事务中调用外部 API

    参数:
        db:           数据库会话
        source_id:    来源记录 ID
        evaluator_id: 研判人 ID（逻辑关联 sys_user）
        rule_id:      指定规则 ID（可选，不填则自动按优先级匹配）

    返回:
        {source_id: int, parse_status: str}
    """
    # 1. 校验来源记录存在
    source = db.query(CustomerSource).filter_by(id=source_id).first()
    if not source:
        raise NotFoundError(f"客户来源记录不存在: id={source_id}")

    # 2. 状态校验：只有 pending 或 failed 状态可以触发研判
    if source.parse_status == "success":
        raise StateError("该客户资料已研判成功，无需重复研判")
    # 如果正在处理中，允许重试（将状态重置为 pending）

    # 3. ⭐ 逻辑外键校验：研判人存在
    if evaluator_id is not None:
        evaluator = db.query(SysUser).filter_by(id=evaluator_id).first()
        if not evaluator:
            raise ReferenceNotFoundError("研判人", evaluator_id)

    # 4. ⭐ 逻辑外键校验：指定规则存在
    if rule_id is not None:
        rule = db.query(ProfileRule).filter_by(id=rule_id).first()
        if not rule:
            raise ReferenceNotFoundError("研判规则", rule_id)

    # 5. 重置解析状态为 pending（在请求事务中完成）
    source.parse_status = "pending"
    source.parse_error = None  # 清除之前的失败原因
    db.commit()

    # 6. 提交后台异步任务（使用独立 Session）
    # 注意: 这里只是示意结构，实际需要 FastAPI BackgroundTasks
    # 在 router 层用 background_tasks.add_task(_do_analyze, source_id, evaluator_id, rule_id)
    # 后台任务函数见下方 _do_analyze

    return {"source_id": source_id, "parse_status": "pending"}


def execute_analysis(
    source_id: int,
    evaluator_id: Optional[int] = None,
    rule_id: Optional[int] = None,
) -> CustomerProfile:
    """
    执行 AI 研判（由后台异步任务调用，使用独立 Session）。

    ⚠️ 此函数在 BackgroundTasks 中执行，不在请求事务中。

    流程:
      1. 独立 Session 查询来源记录
      2. 提取客户数据（从 parse_result 或 raw_content）
      3. 从 .md 文件加载规则 → 调用 LLM 语义研判（事务外）
      4. 写入 customer_profile（独立事务）
      5. 更新 customer_source.parse_status

    参数:
        source_id:    来源记录 ID
        evaluator_id: 研判人 ID
        rule_id:      指定规则 ID（保留兼容，当前不使用 DB 规则）
    """
    db = SessionLocal()  # ← 独立 Session
    try:
        # 1. 查询来源记录
        source = db.query(CustomerSource).filter_by(id=source_id).first()
        if not source:
            logger.error(f"研判失败: 来源记录不存在 source_id={source_id}")
            return None

        # 2. 准备客户数据（从 parse_result 或 raw_content 提取）
        customer_data = _extract_customer_data(source)

        # 3. 调用 LLM 进行语义研判（事务外调用外部 API）
        # ============================================
        # 规则从 .md 文件动态加载，LLM 根据规则全文做语义匹配。
        # 规则不在代码中——修改 .md 文件即可调整规则。
        # ============================================
        try:
            ai_result = _llm_match_analysis(
                customer_data=customer_data,
                content_text=source.raw_content or "",
            )
        except Exception as e:
            logger.exception(f"LLM研判异常 source_id={source_id}")
            source.parse_status = "failed"
            source.parse_error = str(e)[:500]
            db.commit()
            return None
        # ============================================

        # 5. 保存研判结果（独立事务）
        profile = CustomerProfile(
            customer_name=customer_data.get("name") or source.raw_content[:64] if source.raw_content else None,
            contact_info=customer_data.get("contact_info"),
            source_id=source_id,
            background_info=customer_data,
            match_result=ai_result.get("match_result"),
            matched_product=ai_result.get("matched_product"),
            match_score=ai_result.get("match_score"),
            match_reason=ai_result.get("match_reason"),
            recommended_programs=ai_result.get("recommended_programs"),
            evaluator_id=evaluator_id,
        )
        db.add(profile)

        # 更新来源记录状态
        source.parse_status = "success"
        source.parse_result = customer_data
        db.commit()

        db.refresh(profile)
        return profile

    except Exception as e:
        logger.exception(f"研判失败 source_id={source_id}: {str(e)}")
        # 标记失败
        try:
            source = db.query(CustomerSource).filter_by(id=source_id).first()
            if source:
                source.parse_status = "failed"
                source.parse_error = str(e)[:500]  # 截断，防止超长
                db.commit()
        except Exception:
            logger.exception("更新失败状态时出错")
        return None
    finally:
        db.close()


def _extract_file_content(file_path: str, ext: str) -> Optional[str]:
    """
    从上传的文件中提取文本内容（对应需求 CR-003）。

    支持格式:
      .pdf  → PyPDF2 逐页提取文本
      .xlsx → openpyxl 逐行拼接（制表符分隔）
      .txt  → 直接读取文件内容
      .docx → 暂不支持（返回 None）

    解析失败时返回 None 而非抛异常——文件内容提取是增强功能，
    不应阻塞主上传流程。调用方会 fallback 到 raw_content 文本。

    参数:
        file_path: 文件的绝对路径
        ext:       文件扩展名（含点，如 ".pdf"）

    返回:
        提取的文本内容，失败时返回 None
    """
    try:
        if ext == ".pdf":
            from PyPDF2 import PdfReader
            reader = PdfReader(file_path)
            pages = []
            for page in reader.pages[:10]:  # 最多提取前 10 页
                text = page.extract_text()
                if text:
                    pages.append(text.strip())
            return "\n".join(pages) if pages else None

        elif ext in (".xlsx", ".xls"):
            from openpyxl import load_workbook
            wb = load_workbook(file_path, read_only=True, data_only=True)
            rows = []
            for sheet in wb.worksheets[:3]:  # 最多 3 个工作表
                for row in sheet.iter_rows(values_only=True):
                    row_text = "\t".join(str(cell) if cell is not None else "" for cell in row)
                    if row_text.strip():
                        rows.append(row_text)
                if len(rows) > 500:  # 最多 500 行，防止超大 Excel
                    break
            wb.close()
            return "\n".join(rows) if rows else None

        elif ext == ".txt":
            with open(file_path, "r", encoding="utf-8", errors="ignore") as f:
                return f.read(10000)  # 最多读取 1 万字符

        else:
            logger.debug(f"不支持的文件类型: {ext}")
            return None

    except Exception as e:
        logger.warning(f"文件内容提取失败 ext={ext}: {str(e)}")
        return None


def _extract_customer_data(source: CustomerSource) -> Dict[str, Any]:
    """
    从来源记录中提取客户数据。

    如果已有 AI 解析结果（parse_result），直接使用；
    否则使用原始文本（raw_content）作为 name 字段。
    """
    if source.parse_result and isinstance(source.parse_result, dict):
        return dict(source.parse_result)  # 返回副本，避免意外修改

    # 原始文本 → 简单提取
    return {
        "name": source.raw_content[:64] if source.raw_content else None,
        "raw_text": source.raw_content,
        "source_type": source.source_type,
        "file_name": source.file_name,
    }


def _load_rules_context() -> str:
    """
    从 .md 文件加载产品线规则 + 全产品目录，拼接为 LLM 上下文。

    规则不在代码中——修改 .md 文件即可更新规则，无需改代码。
    """
    parts = []
    for label, path in [
        ("产品线匹配规则", PRODUCT_RULES_PATH),
        ("全产品线目录", PRODUCT_CATALOG_PATH),
    ]:
        if os.path.exists(path):
            with open(path, "r", encoding="utf-8") as f:
                text = f.read()
            parts.append(text)
            logger.info(f"已加载{label}: {path} ({len(text)}字符)")
        else:
            logger.warning(f"{label}文件不存在: {path}")
    return "\n\n".join(parts)


def _llm_match_analysis(
    customer_data: Dict[str, Any],
    content_text: str = "",
) -> Dict[str, Any]:
    """
    调用 LLM 进行客户-产品线智能语义匹配。

    规则从 .md 文件动态加载，作为 System Prompt 上下文传入 LLM。
    LLM 负责理解规则 → 对客户画像做语义匹配 → 推荐产品线。

    返回字段对齐 customer_profile 表:
      match_result: matched | partial | not_matched
      match_score:  DECIMAL(5,2)
      matched_product / match_reason / recommended_programs(JSON)
    """
    rules_context = _load_rules_context()

    system_prompt = f"""你是一个教育服务机构的资深客户研判专家，精通两大产品线：
- 新加坡国际本硕升学计划（8个项目：2+2定向培养本科班、2+2+1本硕连读、0.5+2国际本科班、0.5+2+1本硕连读、6+6酒店运营大专就业班、9+6航空运营大专就业班、一年制专升本、一年制本升硕）
- 中德精英人才共建计划（16个专业方向：机电一体化技术、工业机械师、精密加工与数控技术、自动化技术、汽车机电一体化、电动汽车与高压系统、车身与车辆制造技术、建筑工程与项目管理、建筑设备与能源技术、金属构造与焊接技术、应用软件开发、IT系统集成与网络安全、综合护理与健康管理、老年康养与康复治疗、酒店与餐饮管理、航空地勤与物流管理）

你的任务是根据客户画像，严格依据以下规则进行语义匹配，推荐最合适的产品线。

---

【产品线匹配规则与全产品目录】

{rules_context}

---

【研判原则（按优先级）】
1. 学历+年龄硬门槛——不满足则不推荐
2. 核心需求对齐——升学/就业/学历提升/移民与产品线定位一致
3. 语言能力评估——英语/德语水平决定是否需要语言培训
4. 专业背景加分——客户专业/工作背景相关时加分
5. 经济能力考量——年收入决定推荐新加坡(自费)还是德国(免学费+补贴)
6. 诚实标注差距——不满足条件在gaps中明确指出

【输出格式】必须只返回合法JSON，首字符为{{，不含markdown标记：

{{
  "customer_name": "姓名或null",
  "background_info": {{"education":"学历","age":年龄或null,"intended_countries":["国家"],"language_scores":{{}},"work_experience":"描述或null","other_info":"其他或null"}},
  "match_result": "matched",
  "matched_product": "最优产品线完整名称(必须与产品目录一字不差)",
  "match_score": 88.5,
  "match_reason": "综合研判理由",
  "candidates": [
    {{
      "name": "产品线完整名称",
      "category": "新加坡/中德",
      "match_level": "★★★",
      "reasons": ["理由1","理由2"],
      "gaps": ["差距1"],
      "cost_note": "费用参考",
      "next_step": "下一步建议"
    }}
  ],
  "match_summary": "一句话概括",
  "customer_snapshot": "学历=X 年龄=Y 核心需求=Z",
  "recommended_programs": [{{"program_name":"项目名","score":90,"reason":"推荐理由"}}]
}}

【约束】
1. product_name必须与产品目录完全一致，不得编造
2. 最多5个candidates，质量优先
3. 条件冲突(如年龄超限)绝不推荐
4. ★★★=学历年龄完全符合+需求对齐+语言达标
5. ★★=学历年龄符合+需求基本对齐 有1-2项差距
6. ★=学历年龄符合但需求不完全对齐 或多项差距
7. match_result: matched=有★★★候选, partial=仅有★★或★, not_matched=无候选
8. match_score: matched≥80, partial=50-79, not_matched<50"""

    customer_json = json.dumps(customer_data, ensure_ascii=False, indent=2)
    user_message = f"""请对以下客户进行产品线精准匹配：

## 客户原始资料
{content_text[:2000] if content_text else "（未提供）"}

## 客户结构化画像
{customer_json}

请按JSON格式输出匹配结果。"""

    from utils.dify_client import call_llm_direct

    logger.info(f"LLM研判开始: customer={customer_data.get('name','unknown')}")
    llm_result = call_llm_direct(
        system_prompt=system_prompt,
        user_message=user_message,
        temperature=0.2,
        response_format="json_object",
    )

    standardized = {
        "customer_name": llm_result.get("customer_name") or customer_data.get("name"),
        "background_info": llm_result.get("background_info") or customer_data,
        "match_result": llm_result.get("match_result", "not_matched"),
        "matched_product": llm_result.get("matched_product"),
        "match_score": llm_result.get("match_score"),
        "match_reason": llm_result.get("match_reason"),
        "candidates": llm_result.get("candidates", []),
        "match_summary": llm_result.get("match_summary"),
        "customer_snapshot": llm_result.get("customer_snapshot"),
        "recommended_programs": llm_result.get("recommended_programs", []),
    }

    logger.info(
        f"LLM研判完成: match_result={standardized['match_result']}, "
        f"score={standardized['match_score']}, candidates={len(standardized['candidates'])}"
    )
    return standardized


# ============================================================
# 五、画像规则管理
# ============================================================


def list_profile_rules(
    db: Session,
    product_line: Optional[str] = None,
    status: Optional[int] = None,
) -> List[ProfileRuleResponse]:
    """
    查询画像规则列表。

    不分页（规则数量通常较少，10-50 条）。

    参数:
        db:           数据库会话
        product_line: 按产品线筛选（可选）
        status:       按状态筛选（可选，默认查所有）

    返回:
        规则列表
    """
    query = db.query(ProfileRule)

    if product_line:
        query = query.filter(ProfileRule.product_line == product_line)
    if status is not None:
        query = query.filter(ProfileRule.status == status)

    rules = query.order_by(ProfileRule.priority.desc(), ProfileRule.id).all()
    return [ProfileRuleResponse.model_validate(r) for r in rules]


def create_profile_rule(db: Session, data: ProfileRuleCreate) -> ProfileRule:
    """
    创建画像研判规则。

    参数:
        db:   数据库会话
        data: 规则创建请求体

    返回:
        新创建的 ProfileRule ORM 对象
    """
    rule = ProfileRule(
        product_line=data.product_line,
        rule_name=data.rule_name,
        rule_content=data.rule_content,
        match_prompt=data.match_prompt,
        priority=data.priority,
        status=1,  # 新建规则默认启用
    )
    db.add(rule)
    db.commit()
    db.refresh(rule)
    return rule


def update_profile_rule(db: Session, rule_id: int, data: ProfileRuleUpdate) -> ProfileRule:
    """
    更新画像研判规则（部分更新）。

    参数:
        db:      数据库会话
        rule_id: 规则 ID
        data:    规则更新请求体（只含需修改的字段）

    返回:
        更新后的 ProfileRule ORM 对象
    """
    # 1. 查规则
    rule = db.query(ProfileRule).filter_by(id=rule_id).first()
    if not rule:
        raise NotFoundError(f"画像规则不存在: id={rule_id}")

    # 2. 只更新非 None 字段
    update_data = data.model_dump(exclude_unset=True)
    for field, value in update_data.items():
        setattr(rule, field, value)

    # 3. 提交
    db.commit()
    db.refresh(rule)
    return rule
